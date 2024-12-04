import openpyxl

def cargar_hoja(ruta, hoja):
    wb = openpyxl.load_workbook(ruta)
    ws = wb[hoja]
    return wb, ws


ruta_pelis = "ficheros/pelis.xlsx"
hoja_pelis = "pelis"
(wb, ws) = cargar_hoja(ruta_pelis, hoja_pelis)
print(ws["B1"].value)
print(ws.max_row)
print(ws.max_column)
print(len(ws[1]))


for fila in ws.values:
    print(fila[0])



ws[1][1].value = "Director"
ws["B1"] = "Director"
ws["D1"] = "Comentarios"
print(ws.max_row)
print(ws.max_column)
print(ws["D1"].value)

nueva_fila = ["Torrente", "Santiago Segura", 1998, "Nueva"]
ws.append(nueva_fila)


print(ws.max_row)
print(ws.max_column)

ruta_renombrada = "ficheros/pelis1.xlsx"
wb.save(ruta_renombrada)